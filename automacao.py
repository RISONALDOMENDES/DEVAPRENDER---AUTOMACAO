'''
1-entrar na planilha e extrair o cpf do cliente
2-entrar no site https://consultcpf-devaprender.netlify.app e usar o cpf da planilha para verificar o status de pagamento de cada cliente
3-verificar se esta "em dia" ou " atrasado"
4- se estiver "em dia" pegar a data do pagamento e metodo de pagamento
5- caso contrario colocar op status como pendente
6- inserir essas novas informações :(nome, cpf, vencimento, status e caso esteja em dia, data do pagamento, metodo do pagamento(cartão ou boleto)) em uma nova planilha
7- repetir o processo ate chegar no ultimo clente
'''

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
import pandas as pd

# 1-entrar na planilha e extrair o cpf do cliente
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    # 2-entrar no site https://consultcpf-devapnrender.netlify.app e usar o cpf da planilha para verificar o status de pagamento de cada cliente
    
    sleep(5)
    campo_pesquisa = driver.find_element(By.XPATH, "//input[@id='cpfInput']")
    sleep(3)
    campo_pesquisa.send_keys(cpf)
    sleep(4)
    # 3-verificar se esta "em dia" ou " atrasado"
    botao_pesquisar = driver.find_element(By.XPATH,"//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(3)
    botao_pesquisar.click()
    sleep(5)
    
    # 4- se estiver "em dia" pegar a data do pagamento e metodo de pagamento
    status = driver.find_element(By.XPATH, "//span[@id= 'statusLabel']")
    if status.text == 'em dia':
        data_pagamento = driver.find_eleent(By.XPATH, "//p[@id= 'paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH, "//p[@id= 'paymentMethod']")
        
        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[4
                                                               ]
             
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia',  data_pagamento_limpo, metodo_pagamento_limpo])
    else:
        # 5- caso contrario colocar op status como pendente
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Sheet1']
        
        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
    # 6- inserir essas novas informações :(nome, cpf, vencimento, status e caso esteja em dia, data do pagamento, metodo do pagamento(cartão ou boleto)) em uma nova planilha
    # 7- repetir o processo ate chegar no ultimo clente
    